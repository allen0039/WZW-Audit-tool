#!/usr/bin/env python3
"""完整重新生成题库"""
import os, re, json, shutil
from collections import OrderedDict
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR if (SCRIPT_DIR / 'tiku').exists() else SCRIPT_DIR.parent
BASE = PROJECT_ROOT / 'tiku' / '北京优度'
OUT = PROJECT_ROOT
SCRIPT = OUT / 'shenjitools.user.js'
RELEASE_DIR = OUT / 'release'
LEGACY_SCRIPT_NAMES = [
    '审计实训万能填表助手.user.js',
    '审计实训万能填表助手v6.user.js',
    '审计实训万能填表助手v7.user.js',
]

INDEX_PAT = re.compile(r'^[A-Z]{1,4}\d*(?:-\d+)*$')
PRODUCT_PAT = re.compile(r'^[A-Z]\d{2}$')
SINGLE_PAT = re.compile(r'^[A-Z]$')
KNOWN_XREF = {'AB','AC','AD','AE','BC','BD','BE','CD','CE','DE','FY','GF','GJ','RK'}

def scan_excel():
    import openpyxl
    files = sorted(os.path.join(root, f) for root, _, fs in os.walk(BASE) for f in fs 
                   if f.endswith('.xlsx') and not f.startswith('~$'))
    result = OrderedDict()
    
    for fp in files:
        rel = os.path.relpath(fp, BASE)
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
        except:
            continue
        for sn in (s for s in wb.sheetnames if not (s.lower().startswith('sheet') and s.lower() != 'sheet1')):
            ws = wb[sn]
            max_r = min(ws.max_row or 800, 800)
            max_c = min(ws.max_column or 25, 25)
            
            # Read all data into memory
            rows = []
            for rr in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c, values_only=True):
                rd = [(ci, str(v).strip()) for ci, v in enumerate(rr) if v is not None and str(v).strip()]
                rows.append(rd)
            
            # Find table indexes
            indexes = OrderedDict()
            for ri, rd in enumerate(rows):
                for ci, val in rd:
                    if not INDEX_PAT.match(val) or PRODUCT_PAT.match(val) or (SINGLE_PAT.match(val) and len(val)==1):
                        continue
                    is_idx = ri <= 5 and 5 <= ci <= 15
                    if not is_idx:
                        for ci2, val2 in rd:
                            if ci2 == ci-1 and '索引号' in val2:
                                is_idx = True
                                break
                        is_idx = is_idx or (ri <= 10 and '索引号' in str(rows[ri-1]) if ri > 0 else False)
                    if is_idx and val not in indexes:
                        indexes[val] = (ri, ci)
            
            for idx, (sr, sc) in sorted(indexes.items(), key=lambda x: x[1][0]):
                if idx in result:
                    continue
                # Extract table
                cells = []
                end_r = None
                blank = 0
                for i in range(sr, len(rows)):
                    if not rows[i]:
                        blank += 1
                        if blank >= 2:
                            for j in range(i+1, len(rows)):
                                if rows[j]:
                                    for c_, v_ in rows[j]:
                                        if (INDEX_PAT.match(v_) and not PRODUCT_PAT.match(v_) and 
                                            not SINGLE_PAT.match(v_) and 7 <= c_ <= 12 and v_ not in KNOWN_XREF):
                                            end_r = i - blank + 1
                                            break
                                    break
                            if end_r is None and blank >= 3:
                                end_r = i - blank + 1
                            if end_r:
                                break
                    else:
                        blank = 0
                if end_r is None:
                    end_r = len(rows)
                for i in range(sr, end_r):
                    for ci, val in rows[i]:
                        cells.append([i-sr, ci, val])
                if len(cells) >= 5:
                    title = next((v for _, ci, v in cells if ci == 1), '')
                    result[idx] = {'index_no': idx, 'title': title, 'file': rel, 'sheet': sn,
                                   'cells': [{'r':c[0],'c':c[1],'v':c[2]} for c in cells]}
        wb.close()
    return result

def main():
    extracted = scan_excel()
    print(f'扫描到 {len(extracted)} 条', flush=True)
    
    existing = {}
    try:
        with open(OUT / 'tiku_all.json', encoding='utf-8') as f:
            existing = json.load(f)
    except:
        pass
    print(f'现有 {len(existing)} 条', flush=True)
    
    merged = OrderedDict()
    for k in sorted(existing):
        if k not in extracted:
            merged[k] = existing[k]
    for k in sorted(extracted):
        merged[k] = extracted[k]
    print(f'合并: {len(merged)} 条', flush=True)
    
    # Write tiku_all.json
    with open(OUT / 'tiku_all.json', 'w', encoding='utf-8') as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print('tiku_all.json 完成')
    
    # Write tiku_min.json
    tmin = OrderedDict(
        (k, [v.get('title', ''), [[c['r'], c['c'], c['v']] for c in v['cells']]])
        for k, v in merged.items()
        if v.get('cells') and not k.startswith('_table_')
    )
    with open(OUT / 'tiku_min.json', 'w', encoding='utf-8') as f:
        json.dump(tmin, f, ensure_ascii=False)
    print(f'tiku_min.json 完成 ({len(tmin)} 条)')
    
    # Write tiku_data.js
    parts = ['// 审计实训万能填表助手 - 题库数据', '// 请与 user.js 脚本放在同一目录', '', 'const TIKU_DATA = {']
    for i, (k, v) in enumerate(tmin.items()):
        sep = ',' if i < len(tmin)-1 else ''
        parts.append(f'  "{k}": {json.dumps(v, ensure_ascii=False)}{sep}')
    parts.append('};')
    parts.append('if (typeof module !== "undefined" && module.exports) module.exports = TIKU_DATA;')
    with open(OUT / 'tiku_data.js', 'w', encoding='utf-8') as f:
        f.write('\n'.join(parts))
    print('tiku_data.js 完成')
    
    # Update script
    with open(SCRIPT, 'r', encoding='utf-8') as f:
        script = f.read()
    
    # Build TIKU string
    tiku_str = '{'
    for i, (k, v) in enumerate(tmin.items()):
        sep = ',' if i > 0 else ''
        tiku_str += f'{sep}"{k}":{json.dumps(v, ensure_ascii=False)}'
    tiku_str += '}'
    
    script_new = re.sub(r'const TIKU=\{.*?\};', f'const TIKU={tiku_str};', script, flags=re.DOTALL)
    script_new = re.sub(r'@name\s+[^\n]+', '@name         shenjitools', script_new)
    script_new = re.sub(r'@version\s+\S+', '@version      7.1', script_new)
    script_new = re.sub(r'@description\s+[^\n]+', '@description  审计实训万能填表助手 v7.1 - 完整309条题库', script_new)
    script_new = re.sub(r'万能填表就绪\(\d+题库\)', f'万能填表就绪({len(tmin)}题库)', script_new)
    
    with open(SCRIPT, 'w', encoding='utf-8') as f:
        f.write(script_new)
    print(f'{SCRIPT.name} 完成')

    if RELEASE_DIR.exists():
        shutil.copyfile(OUT / 'tiku_min.json', RELEASE_DIR / 'tiku_min.json')
        shutil.copyfile(OUT / 'tiku_data.js', RELEASE_DIR / 'tiku_data.js')
        shutil.copyfile(SCRIPT, RELEASE_DIR / 'shenjitools.user.js')
        for name in LEGACY_SCRIPT_NAMES:
            legacy = RELEASE_DIR / name
            if legacy.exists():
                legacy.unlink()
        print('release 文件同步完成')

    for name in LEGACY_SCRIPT_NAMES:
        legacy = OUT / name
        if legacy.exists():
            legacy.unlink()
    
    total = len(merged)
    new_k = sorted(set(extracted) - set(existing))
    old_k = sorted(set(existing) - set(extracted))
    print(f'\n✅ 完成! 总共 {total} 条')
    print(f'   新增: {len(new_k)} 条: {", ".join(new_k)}')
    print(f'   保留旧: {len(old_k)} 条(未重新扫描到)')

if __name__ == '__main__':
    main()
